#!/usr/bin/env python3
"""
generate_projects_json.py
Run by GitHub Action after sync_data.yml to produce projects.json.gz.
Pre-aggregates 825K DN dump rows → ~32K project-level rows (~1.2 MB gz vs 14.7 MB)

Metering is calculated per-project using the backend rate table formula:
  metering = NM_rate(city, phase) + GM_rate(city, phase) + DN_dump_metering_items
No manual monthly totals needed — fully automatic for any month.

COGS uses the ERP Categorization file as the PRIMARY source for item_code → category
mapping, falling back to the raw DN item_category only when the item_code is not in
the ERP file.  Safety Lifeline and Civil Work are excluded from COGS.

Usage:  python3 generate_projects_json.py
Input:  data.csv.gz  (same directory)
        erp_categorization.csv (same directory — optional, uses embedded map if missing)
Output: projects.json.gz (same directory)
"""
import gzip, csv, io, json, os, re, sys
from collections import defaultdict
from datetime import datetime

# ── Cohort assignment (QCD-based) ────────────────────────────────────────────
GZ_OFFERS_SET = {'GoodZero', 'GoodZero Pro', 'GoodZero Uno', 'GoodZero+'}

# ── Configuration ────────────────────────────────────────────────────────────
CIVL_TO_ELEC   = {'CIVL-0012','CIVL-0013','CIVL-0014','CIVL-0015','CIVL-0016'}
METERING_REMAP = {'ACDB-2449-EATON'}
DONGLE_PFX     = {'DALO','DALA'}

COGS_CATS = {
    'Module','Inverter','Prefab MMS','Cables','I&C KIT','Conduit Pipe',
    'Earthing & LA','Junction Box','Tin Shed MMS','Safety','I&C Accessories',
    'Welded MMS','SS NBW','Electrical BoS','Data Logger','Metering','Welcome Kit and Board','Ladder'
}

EXCLUDE_CATS = {'Safety Lifeline', 'Civil Work', 'Civil work'}

NM_RATES = {
    'Pune':        (0, 0),
    'Nashik':      (0, 0),
    'Nagpur':      (0, 0),
    'Aurangabad':  (0, 0),
    'Jalgaon':     (0, 0),
    'Ahmednagar':  (0, 0),
    'Latur':       (0, 0),
    'Kolhapur':    (0, 0),
    'Mumbai':      (0, 0),
    'Amravati':    (0, 0),
    'Solapur':     (0, 0),
    'Bhopal':      (2841, 4617),
    'Indore':      (6800, 9050),
    'Jabalpur':    (9785, 14050),
    'Gwalior':     (2841, 4617),
    'Bengaluru':   (3250, 6376),
    'Hyderabad':   (0, 0),
    'Ahmedabad':   (0, 0),
    'Surat':       (0, 0),
    'Baroda':      (0, 0),
    'Jaipur':      (3550, 6650),
    'Ajmer':       (3550, 6650),
    'Kota':        (3550, 6650),
    'Lucknow':     (1350, 4350),
    'Kanpur':      (1350, 4350),
    'Varanasi':    (1350, 4350),
    'Noida':       (1350, 4350),
    'NCR':         (0, 0),
    'Meerut':      (1350, 4350),
    'Bareilly':    (1350, 4350),
    'Kochi':       (3250, 6376),
    'Chennai':     (2763, 5011),
    'Agra':        (1350, 4350),
    'Coimbatore':  (2763, 5011),
    'Salem':       (2763, 5011),
    'Raipur':      (0, 0),
    'Mysuru':      (3250, 6376),
    'Warangal':    (0, 0),
    'Gurgaon':     (0, 0),
    'Delhi NCR':   (0, 0),
    'Ghaziabad':   (1350, 4350),
    'Vijayawada':  (0, 0),
}

GM_RATES = {
    'Pune':        (1260, 2620),
    'Nashik':      (1260, 2620),
    'Nagpur':      (1260, 2620),
    'Aurangabad':  (1260, 2620),
    'Jalgaon':     (1260, 2620),
    'Ahmednagar':  (1260, 2620),
    'Latur':       (1260, 2620),
    'Kolhapur':    (1260, 2620),
    'Mumbai':      (1260, 2620),
    'Amravati':    (1260, 2620),
    'Solapur':     (1260, 2620),
    'Bhopal':      (0, 0),
    'Indore':      (0, 0),
    'Jabalpur':    (0, 0),
    'Gwalior':     (0, 0),
    'Bengaluru':   (0, 0),
    'Hyderabad':   (0, 0),
    'Ahmedabad':   (0, 0),
    'Surat':       (0, 0),
    'Baroda':      (0, 0),
    'Jaipur':      (3050, 5650),
    'Ajmer':       (3050, 5650),
    'Kota':        (3050, 5650),
    'Lucknow':     (0, 0),
    'Kanpur':      (0, 0),
    'Varanasi':    (0, 0),
    'Noida':       (0, 0),
    'NCR':         (0, 0),
    'Meerut':      (0, 0),
    'Bareilly':    (0, 0),
    'Kochi':       (0, 0),
    'Chennai':     (0, 0),
    'Agra':        (0, 0),
    'Coimbatore':  (0, 0),
    'Salem':       (0, 0),
    'Raipur':      (0, 0),
    'Mysuru':      (0, 0),
    'Warangal':    (0, 0),
    'Gurgaon':     (0, 0),
    'Delhi NCR':   (0, 0),
    'Ghaziabad':   (0, 0),
}

def detect_inverter_phase(inv_item_name):
    n = str(inv_item_name)
    if 'Battery' in n and 'Hybrid' in n:
        return None
    if '3 Phase' in n or '3-Phase' in n or 'Three Phase' in n:
        return 'Three Phase'
    if '1 Phase' in n or '1-Phase' in n or 'Single Phase' in n:
        return 'Single Phase'
    if 'ENPHASE' in n.upper() or 'Micro' in n.lower():
        return 'Single Phase'
    return 'Single Phase'

def calc_metering_backend(city, inv_phase, sanction_phase):
    phase_for_nm = sanction_phase if sanction_phase else inv_phase
    nm_idx = 0 if (not phase_for_nm or 'single' in phase_for_nm.lower()) else 1
    nm = NM_RATES.get(city, (0, 0))
    gm_idx = 0 if (not sanction_phase or 'single' in sanction_phase.lower()) else 1
    gm = GM_RATES.get(city, (0, 0))
    return nm[nm_idx] + gm[gm_idx]

def is_metering_dn_item(item_name):
    if 'Communication Modem' in item_name and 'Optical Cable' in item_name:
        return True
    if 'FRP Meter Box' in item_name:
        return True
    if 'Meter Box' in item_name and '400x300x150' in item_name and 'SPARK' in item_name:
        return True
    return False

CELL_CITY_STATE = {
    'Aurangabad Expansion':{'c':'Aurangabad','s':'MH East'},
    'Bangalore Royal Challengers':{'c':'Bengaluru','s':'Karnataka'},
    'Bangalore Royal Challengers**':{'c':'Bengaluru','s':'Karnataka'},
    'Bangalore Royal Challengers 2':{'c':'Bengaluru','s':'Karnataka'},
    'Bengaluru Royal Challengers':{'c':'Bengaluru','s':'Karnataka'},
    'Bengaluru Royal Challengers 2':{'c':'Bengaluru','s':'Karnataka'},
    'Baroda Blasters':{'c':'Baroda','s':'Gujrat'},
    'Baroda Smashers':{'c':'Baroda','s':'Gujrat'},
    'Bhopal Strikers':{'c':'Bhopal','s':'Madhya Pradesh'},
    'Bhopal Strikers 2':{'c':'Bhopal','s':'Madhya Pradesh'},
    'Bhopal Strikers 3':{'c':'Bhopal','s':'Madhya Pradesh'},
    'Bhopal Strikers 4':{'c':'Bhopal','s':'Madhya Pradesh'},
    'Bhopal Strikers 5':{'c':'Bhopal','s':'Madhya Pradesh'},
    'Bhopal Strikers 6':{'c':'Bhopal','s':'Madhya Pradesh'},
    'Bhopal Strikers**':{'c':'Bhopal','s':'Madhya Pradesh'},
    'Delhi Dashers 2':{'c':'Gurgaon','s':'Delhi'},
    'Delhi Dashers 3':{'c':'Delhi NCR','s':'Delhi'},
    'Delhi Dashers 4':{'c':'Ghaziabad','s':'Delhi'},
    'Delhi Dashers 5':{'c':'Delhi NCR','s':'Delhi'},
    'Delhi Dashers 6':{'c':'Delhi NCR','s':'Delhi'},
    'Gujrat Gladiators':{'c':'Ahmedabad','s':'Gujrat'},
    'Gujrat Gladiators 2':{'c':'Ahmedabad','s':'Gujrat'},
    'Gujarat Gladiators':{'c':'Ahmedabad','s':'Gujrat'},
    'Gujarat Gladiators 2':{'c':'Ahmedabad','s':'Gujrat'},
    'Ahmedabad Gladiators':{'c':'Ahmedabad','s':'Gujrat'},
    'Ahmedabad Gladiators 2':{'c':'Ahmedabad','s':'Gujrat'},
    'Gwalior Groundbreakers':{'c':'Gwalior','s':'Madhya Pradesh'},
    'Gwalior Groundbreakers 2':{'c':'Gwalior','s':'Madhya Pradesh'},
    'Gwalior Groundbreakers 3':{'c':'Gwalior','s':'Madhya Pradesh'},
    'Gwalior Groundbreakers 4':{'c':'Gwalior','s':'Madhya Pradesh'},
    'Gwalior Groundbreakers 5':{'c':'Gwalior','s':'Madhya Pradesh'},
    'Speed Order Gwalior 5':{'c':'Gwalior','s':'Madhya Pradesh'},
    'Indore Immortals':{'c':'Indore','s':'Madhya Pradesh'},
    'Indore Immortals 2':{'c':'Indore','s':'Madhya Pradesh'},
    'Indore Immortals 3':{'c':'Indore','s':'Madhya Pradesh'},
    'Indore Immortals 4':{'c':'Indore','s':'Madhya Pradesh'},
    'Indore Immortals 5':{'c':'Indore','s':'Madhya Pradesh'},
    'Indore Immortals 6':{'c':'Indore','s':'Madhya Pradesh'},
    'Indore Immortals 7':{'c':'Indore','s':'Madhya Pradesh'},
    'Jabalpur Champions':{'c':'Jabalpur','s':'Madhya Pradesh'},
    'Jabalpur Champions 2':{'c':'Jabalpur','s':'Madhya Pradesh'},
    'Jabalpur Champions 3':{'c':'Jabalpur','s':'Madhya Pradesh'},
    'Jabalpur Champions 4':{'c':'Jabalpur','s':'Madhya Pradesh'},
    'Jabalpur Champions 5':{'c':'Jabalpur','s':'Madhya Pradesh'},
    'Jalgaon Expansion':{'c':'Jalgaon','s':'MH East'},
    'Jalgaon Expansion 2':{'c':'Jalgaon','s':'MH East'},
    'Kolhapur Kings':{'c':'Kolhapur','s':'MH West'},
    'Lucknow Lions':{'c':'Lucknow','s':'Uttar Pradesh'},
    'Lucknow Lions 2':{'c':'Lucknow','s':'Uttar Pradesh'},
    'Lucknow Lions 3':{'c':'Lucknow','s':'Uttar Pradesh'},
    'Lucknow Lions 4':{'c':'Lucknow','s':'Uttar Pradesh'},
    'Speed Order Lucknow 4':{'c':'Lucknow','s':'Uttar Pradesh'},
    'Speed Order Lucknow 5':{'c':'Lucknow','s':'Uttar Pradesh'},
    'Noida Knight Riders':{'c':'Noida','s':'Uttar Pradesh'},
    'Kanpur Tigers':{'c':'Kanpur','s':'Uttar Pradesh'},
    'Kanpur Tigers 2':{'c':'Kanpur','s':'Uttar Pradesh'},
    'Kanpur Tigers 3':{'c':'Kanpur','s':'Uttar Pradesh'},
    'Varanasi Warriors':{'c':'Varanasi','s':'Uttar Pradesh'},
    'Agra Knight Riders':{'c':'Agra','s':'Uttar Pradesh'},
    'Nagpur Daredevils':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 2':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 3':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 4':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 5':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 6':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 7':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 8':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 9':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 10':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 13':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 14':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 15':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils Temp':{'c':'Nagpur','s':'MH East'},
    'Amravati Riders':{'c':'Amravati','s':'MH East'},
    'Amravati Riders 3':{'c':'Amravati','s':'MH East'},
    'Nashik Finishers':{'c':'Nashik','s':'MH West'},
    'Nashik Finishers 2':{'c':'Nashik','s':'MH West'},
    'Nashik Finishers 3':{'c':'Nashik','s':'MH West'},
    'Nashik Finishers 5':{'c':'Nashik','s':'MH West'},
    'Pune Squadrons':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 2':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 3':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 4':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 5':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 6':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 7':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 8':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 9':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 10':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 11':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 12':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 13':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 14':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 15':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 16':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 17':{'c':'Pune','s':'MH West'},
    'Pune Squadrons Temp':{'c':'Pune','s':'MH West'},
    'Pune Squadrons**':{'c':'Pune','s':'MH West'},
    'Ahilyanagar Regiments':{'c':'Pune','s':'MH West'},
    'Speed Order Ahilyanagar 1':{'c':'Pune','s':'MH West'},
    'Speed Order Pune 11':{'c':'Pune','s':'MH West'},
    'Solapur Super Kings':{'c':'Solapur','s':'MH West'},
    'Surat Expansion':{'c':'Surat','s':'Gujrat'},
    'Surat Expansion 2':{'c':'Surat','s':'Gujrat'},
    'Jaipur Titans':{'c':'Jaipur','s':'Rajasthan'},
    'Speed Order Jaipur 2':{'c':'Jaipur','s':'Rajasthan'},
    'Kota Knights':{'c':'Kota','s':'Rajasthan'},
    'Ajmer Aces':{'c':'Ajmer','s':'Rajasthan'},
    'Ajmer Aces 2':{'c':'Ajmer','s':'Rajasthan'},
    'Telangana Tuskers':{'c':'Hyderabad','s':'Telangana'},
    'Telangana Tuskers 2':{'c':'Hyderabad','s':'Telangana'},
    'Telangana Tuskers 3':{'c':'Hyderabad','s':'Telangana'},
    'Hyderabad Tuskers':{'c':'Hyderabad','s':'Telangana'},
    'Hyderabad Tuskers 2':{'c':'Hyderabad','s':'Telangana'},
    'Hyderabad Tuskers 3':{'c':'Hyderabad','s':'Telangana'},
    'Warangal Waveriders':{'c':'Warangal','s':'Telangana'},
    'Vijayawada Strikers':{'c':'Vijayawada','s':'Andhra Pradesh'},
    'Kochi Crushers':{'c':'Kochi','s':'Kerala'},
    'Raipur Royals':{'c':'Raipur','s':'Chhattisgarh'},
    'Chennai Super Kings':{'c':'Chennai','s':'Tamil Nadu'},
    'Chennai Super Kings 2':{'c':'Chennai','s':'Tamil Nadu'},
    'Speed Order Chennai 3':{'c':'Chennai','s':'Tamil Nadu'},
    'Coimbatore Kovai Kings':{'c':'Coimbatore','s':'Tamil Nadu'},
    'Mysuru Mavericks':{'c':'Mysuru','s':'Karnataka'},
    'Speed Order Gurgaon':{'c':'Gurgaon','s':'Delhi'},
    'Latur Expansion':{'c':'Latur','s':'MH East'},
    'Ahmednagar Expansion':{'c':'Ahmednagar','s':'MH West'},
    'Noida Thunders':{'c':'Noida','s':'Uttar Pradesh'},
    'Salem Spartans':{'c':'Salem','s':'Tamil Nadu'},
}

def detect_inverter_type(item_name):
    n = str(item_name)
    if 'Battery' in n and 'Hybrid' in n:
        return None
    if 'ENPHASE' in n.upper() or 'Micro' in n.lower():
        return 'Enphase'
    m = re.search(r'(\d+\.?\d*)\s*[kK][wW]', n)
    if not m:
        return 'Other'
    kw = m.group(1)
    try:
        kw_f = float(kw)
        kw = str(int(kw_f)) if kw_f == int(kw_f) else str(kw_f)
    except: pass
    if 'Hybrid' in n:
        return f'{kw} kW Hybrid'
    if '3 Phase' in n or '3-Phase' in n or 'Three Phase' in n:
        return f'{kw} kW 3 Phase'
    return f'{kw} kW'

MON_MAP = {'jan':0,'feb':1,'mar':2,'apr':3,'may':4,'jun':5,'jul':6,'aug':7,'sep':8,'oct':9,'nov':10,'dec':11}

def parse_date(v):
    if not v: return None
    parts = v.strip().split('-')
    if len(parts) == 3:
        try:
            day = int(parts[0])
            mon = MON_MAP.get(parts[1].lower()[:3])
            yr  = int(parts[2]); yr = 2000+yr if yr < 100 else yr
            if mon is not None: return datetime(yr, mon+1, day)
        except: pass
    try: return datetime.strptime(v.strip(), '%Y-%m-%d')
    except: pass
    try: return datetime.strptime(v.strip(), '%d/%m/%Y')
    except: pass
    try: return datetime.strptime(v.strip(), '%Y/%m/%d')
    except: pass
    return None


def load_pricing_cohorts(filepath='pricing cohorts.xlsx'):
    gz_cohorts   = []
    non_gz_cohorts = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb['Sheet1']
        rows = list(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True))
        for r in rows:
            gz_start, gz_end, gz_name = r[0], r[1], r[2]
            ngz_start, ngz_end, ngz_name = r[3], r[4], r[5]
            if gz_name and gz_start:
                s = gz_start if isinstance(gz_start, datetime) else parse_date(str(gz_start))
                e = gz_end   if isinstance(gz_end,   datetime) else (parse_date(str(gz_end)) if gz_end else None)
                if s: gz_cohorts.append((s, e, str(gz_name).strip()))
            if ngz_name:
                s = ngz_start if isinstance(ngz_start, datetime) else (parse_date(str(ngz_start)) if ngz_start else None)
                e = ngz_end   if isinstance(ngz_end,   datetime) else (parse_date(str(ngz_end))   if ngz_end   else None)
                non_gz_cohorts.append((s, e, str(ngz_name).strip()))
        gz_cohorts.sort(key=lambda x: x[0])
        non_gz_cohorts.sort(key=lambda x: (x[0] or datetime.min))
        print(f"  Loaded {len(gz_cohorts)} GZ cohorts, {len(non_gz_cohorts)} Non-GZ cohorts from {filepath}")
    except ImportError:
        print("  ⚠ openpyxl not installed — cohort assignment disabled. Run: pip install openpyxl")
    except FileNotFoundError:
        print(f"  ⚠ {filepath} not found — cohort assignment disabled.")
    except Exception as ex:
        print(f"  ⚠ Could not load pricing cohorts: {ex}")
    return gz_cohorts, non_gz_cohorts

GZ_COHORTS, NON_GZ_COHORTS = load_pricing_cohorts()


def assign_cohort(qcd_date, offer_type):
    if not qcd_date:
        return ''
    is_gz = offer_type.strip().replace('GoodZero+','GoodZero') in GZ_OFFERS_SET
    cohorts = GZ_COHORTS if is_gz else NON_GZ_COHORTS
    for (start, end, name) in reversed(cohorts):
        if start is None:
            if end and qcd_date <= end:
                return name
        else:
            if qcd_date >= start and (end is None or qcd_date < end):
                return name
    return ''


BOOKING_DUMP_URL = (
    'https://docs.google.com/spreadsheets/d/'
    '1NmE-MH9NyLFcbX1JH--j3yqT32sahvJGj82uFK6l9CY'
    '/gviz/tq?tqx=out:csv&gid=628408580'
)
BOOKING_DUMP_CACHE = 'booking_dump.csv'


def _detect_booking_cols(headers):
    sse_col = qcd_col = offer_col = None
    for h in headers:
        hl = h.strip().lower().replace(' ','').replace('_','').replace('-','').replace('/','')
        if not sse_col and hl in ('sseid','ssid','projectid','projid','sseno','sse'):
            sse_col = h
        if not qcd_col and hl in ('qcd','qcdlqud','lqud','qcddate','quotecompletiondate',
                                   'quoteclosuredate','quotationdate','closuredate',
                                   'quotedate','qcdate','quotationcreationdate'):
            qcd_col = h
        if not offer_col and hl in ('offertype','offeringtype','offer','product',
                                    'producttype','schemetype'):
            offer_col = h
    if not sse_col:
        for h in headers:
            if 'sse' in h.lower(): sse_col = h; break
    if not qcd_col:
        for h in headers:
            hl = h.strip().lower()
            if 'qcd' in hl or 'lqud' in hl: qcd_col = h; break
    return sse_col, qcd_col, offer_col


def _parse_booking_rows(rows_raw):
    qcd_map = {}
    if not rows_raw:
        return qcd_map
    headers = list(rows_raw[0].keys())
    sse_col, qcd_col, offer_col = _detect_booking_cols(headers)
    if not sse_col:
        print(f"    ⚠ SSE ID column not found. Headers: {headers[:10]}")
        return qcd_map
    if not qcd_col:
        print(f"    ⚠ QCD date column not found. Headers: {headers[:15]}")
        return qcd_map
    loaded = 0
    for row in rows_raw:
        sse = row.get(sse_col, '').strip()
        if not sse:
            continue
        qcd_raw = row.get(qcd_col, '')
        qcd_dt = None
        if isinstance(qcd_raw, datetime):
            qcd_dt = qcd_raw
        elif qcd_raw:
            qcd_dt = parse_date(str(qcd_raw).strip())
        offer_raw = row.get(offer_col, '').strip() if offer_col else ''
        qcd_map[sse] = {'qcd': qcd_dt, 'offer': offer_raw}
        loaded += 1
    valid = sum(1 for v in qcd_map.values() if v['qcd'])
    print(f"    Parsed {loaded:,} rows → {valid:,} with valid QCD dates")
    return qcd_map


def load_booking_dump():
    import urllib.request, io

    qcd_map = {}

    print(f"  Fetching booking dump from Google Sheets...")
    fetched_csv = None
    try:
        req = urllib.request.Request(
            BOOKING_DUMP_URL,
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            fetched_csv = resp.read().decode('utf-8', errors='replace')
        print(f"    ✅ Fetched {len(fetched_csv):,} chars from Google Sheets")
    except Exception as ex:
        print(f"    ⚠ Live fetch failed: {ex}")

    if fetched_csv:
        try:
            reader = csv.DictReader(io.StringIO(fetched_csv))
            rows_raw = list(reader)
            if rows_raw:
                qcd_map = _parse_booking_rows(rows_raw)
                if qcd_map:
                    try:
                        with open(BOOKING_DUMP_CACHE, 'w', encoding='utf-8', newline='') as f:
                            writer = csv.DictWriter(f, fieldnames=list(rows_raw[0].keys()))
                            writer.writeheader()
                            writer.writerows(rows_raw)
                        print(f"    💾 Cached {len(rows_raw):,} rows → {BOOKING_DUMP_CACHE}")
                    except Exception as cache_ex:
                        print(f"    ⚠ Could not save cache: {cache_ex}")
                    return qcd_map
            print("    ⚠ Fetched CSV was empty — falling back to cache")
        except Exception as parse_ex:
            print(f"    ⚠ Could not parse fetched CSV: {parse_ex} — falling back to cache")

    if os.path.isfile(BOOKING_DUMP_CACHE):
        print(f"  Reading cached booking dump from {BOOKING_DUMP_CACHE}...")
        try:
            with open(BOOKING_DUMP_CACHE, 'r', encoding='utf-8', errors='replace') as f:
                rows_raw = list(csv.DictReader(f))
            qcd_map = _parse_booking_rows(rows_raw)
            if qcd_map:
                import os as _os
                mtime = datetime.fromtimestamp(_os.path.getmtime(BOOKING_DUMP_CACHE))
                print(f"    ⚠ Using cached data (last updated: {mtime.strftime('%Y-%m-%d %H:%M')})")
                return qcd_map
        except Exception as ex:
            print(f"    ⚠ Could not read cache: {ex}")

    print("  ⚠ Booking dump unavailable — cohort field will be blank for all projects.")
    return qcd_map


print("\nLoading booking dump (QCD dates)...")
QCD_MAP = load_booking_dump()
print(f"  QCD map size: {len(QCD_MAP):,} projects")


ERP_CAT_MAP = {}

def load_erp_categorization():
    global ERP_CAT_MAP
    candidates = [
        'erp_categorization.csv',
        'GMB_GMP_GMI_Mar_26_ERP_Categorization.csv',
        'GMB_GMP_GMI_ERP_Categorization.csv',
        'erp_cat.csv',
    ]
    for fname in candidates:
        if os.path.isfile(fname):
            print(f"Loading ERP categorization from {fname}...")
            with open(fname, 'r', encoding='utf-8', errors='replace') as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames
                code_col = None
                cat_col  = None
                for h in headers:
                    hl = h.strip().lower()
                    if hl in ('item_code', 'item code', 'itemcode', 'code'):
                        code_col = h
                    if hl in ('item_category', 'item category', 'category', 'itemcategory'):
                        cat_col = h
                if not code_col or not cat_col:
                    print(f"  ⚠ Could not find item_code/category columns in {fname}")
                    continue
                count = 0
                for row in reader:
                    ic = row[code_col].strip()
                    ca = row[cat_col].strip()
                    if ic and ca:
                        ERP_CAT_MAP[ic] = ca
                        count += 1
                print(f"  Loaded {count:,} item_code → category mappings")
                return True
    print("  ⚠ No ERP categorization file found — using raw DN categories only")
    return False

erp_loaded = load_erp_categorization()


def resolve_cat(item_code, raw_cat, item_subcategory=''):
    pfx = item_code[:4].upper()
    if pfx in DONGLE_PFX:                     return 'EXCLUDE'
    if item_code in CIVL_TO_ELEC:             return 'Electrical BoS'
    if item_code in METERING_REMAP:           return 'Metering'
    erp = ERP_CAT_MAP.get(item_code)
    if erp:
        if erp in EXCLUDE_CATS:               return 'EXCLUDE'
        if erp == 'Fixtures and Tools':
            if item_subcategory in ('Aluminium Ladder', 'Ladder'):
                return 'Ladder'
            return 'Welcome Kit and Board'
        return erp.strip()
    if raw_cat in EXCLUDE_CATS:               return 'EXCLUDE'
    if raw_cat == 'Fixtures and Tools':
        if item_subcategory in ('Aluminium Ladder', 'Ladder'):
            return 'Ladder'
        return 'Welcome Kit and Board'
    if not raw_cat and item_code.startswith('INVS'): return 'Inverter'
    return raw_cat.strip()


CAT_KEY = {
    'Module':'mod','Inverter':'inv','Prefab MMS':'prf','Cables':'cab','I&C KIT':'ick',
    'Conduit Pipe':'con','Earthing & LA':'ear','Junction Box':'jbx','Tin Shed MMS':'tsh',
    'Safety':'saf','I&C Accessories':'ica','Welded MMS':'wel','SS NBW':'ssn',
    'Electrical BoS':'ebo','Data Logger':'dlg','Metering':'mtr','Welcome Kit and Board':'wkt',
    'Ladder':'lad',
    'MMS':'prf','AC Cable':'cab','DC Cable':'cab','ACDB':'ebo','Walkway':'saf',
    'FRP Walkway':'saf','Cable Tray':'ebo','Marketing':'wkt','Water Pipeline':'con',
    'Inverter Cluster':'inv','Hybrid Inverter':'inv','AMC Accessories':'ica',
    'Spare Breaker Panel':'ebo',
}

def shorten_mms_item_name(name, subcat=''):
    n = name.strip()
    if not n:
        return n
    if 'column' in n.lower() or 'column' in subcat.lower():
        m = re.search(r'\b(\dP)\s+(\d+FT)\b', n, flags=re.IGNORECASE)
        if m:
            return m.group(1) + ' ' + m.group(2)
    n = re.sub(r'^(?:GM\s+Bridge\s+)', '', n, flags=re.IGNORECASE)
    n = re.sub(r'^(?:Galvalume\s+)', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s*-\s*(?:SKU|ITEM|PROD)[-\s]?\w+\s*$', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s*\(\s*(?:SKU|ITEM|PROD)[-\s]?\w+\s*\)\s*$', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s*-\s*Solar\s*Square.*$', '', n, flags=re.IGNORECASE)
    return n.strip()[:80]


def shorten_cable_subcat(name):
    n = name.strip()
    if not n:
        return n
    n = re.sub(r'^(?:Polycab|Havells|RR\s*Kabel|KEI|Finolex|Anchor)\s+', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s+(?:Red|Black|Blue|Green|Yellow|White|Grey|Blue/Black|Red/Black)\s*$', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s*-\s*(?:SKU|ITEM|PROD)[-\s]?\w+\s*$', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s*-\s*Solar\s*Square.*$', '', n, flags=re.IGNORECASE)
    return n.strip()


# ── Build project map ─────────────────────────────────────────────────────────
print("\nReading data.csv.gz...")
project_map = {}
dn_metering = defaultdict(float)
unmapped_cells = defaultdict(int)
unmapped_cats  = defaultdict(int)
excluded_count = 0

proj_inv_types = defaultdict(lambda: defaultdict(lambda: {'qty':0,'amt':0}))
proj_mms_items = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {'qty':0,'amt':0,'uom':''})))
proj_cable_items = defaultdict(lambda: defaultdict(lambda: {'qty':0,'amt':0,'cases':0}))
proj_onm_amt = defaultdict(float)
proj_qhse_amt = defaultdict(float)
erp_mms_overrides = defaultdict(lambda: {'resolved_cat':'','amt':0.0,'rows':0,'sses':set(),'item_name':''})

with gzip.open('data.csv.gz', 'rt', encoding='utf-8', errors='replace') as f:
    reader = csv.DictReader(f)
    for i, row in enumerate(reader):
        if i % 100000 == 0: print(f"  {i:,} rows processed...")
        sse = row['SSE ID'].strip()
        if not sse: continue

        # ── CHANGED: Final Output replaces Final Revenue Excl. GST ──
        rev_raw = (row.get('Final Output', '') or row.get('Final Revenue Excl. GST', '')).strip()
        try: rev = float(rev_raw) if rev_raw else 0
        except: rev = 0

        try: kw = float(row['Project Size (kW)']) if row['Project Size (kW)'].strip() else 0
        except: kw = 0
        try: amt = float(row['amount']) if row['amount'].strip() else 0
        except: amt = 0
        try: qty = float(row['qty']) if row['qty'].strip() else 0
        except: qty = 0
        try: rate = float(row.get('rate','') or 0)
        except: rate = 0

        raw_cat   = row['item_category'].strip()
        item_code = row['item_code'].strip()
        item_name = row['item_name'].strip()
        item_subcat = row.get('item_subcategory', '').strip()
        # ── CHANGED: parent renamed to parent_id ──
        parent_dn = row.get('parent_id', '').strip()
        uom = row.get('uom', '').strip()
        cat = resolve_cat(item_code, raw_cat, item_subcat)

        if raw_cat == 'Prefab MMS' and cat != 'Prefab MMS':
            d = erp_mms_overrides[item_code]
            d['resolved_cat'] = cat
            d['amt'] += amt
            d['rows'] += 1
            d['sses'].add(sse)
            if not d['item_name']: d['item_name'] = item_name

        if parent_dn:
            pl = parent_dn.lower()
            if pl.startswith('onm'):
                proj_onm_amt[sse] += amt
            elif pl.startswith('qhse'):
                proj_qhse_amt[sse] += amt

        if is_metering_dn_item(item_name):
            dn_metering[sse] += amt

        if sse not in project_map:
            # ── CHANGED: Cell Name no longer in new OMS — use .get() safely ──
            cell = row.get('Cell Name', '').strip()
            cs   = CELL_CITY_STATE.get(cell) if cell else None
            city = cs['c'] if cs else row['City'].strip()
            state= cs['s'] if cs else row['State'].strip()
            if cell and not cs and not city:
                unmapped_cells[cell] += 1
            d    = parse_date(row['Installation Completion Date'])
            offer_raw = row['Offer Type'].strip()
            offer= offer_raw.replace('GoodZero+','GoodZero')
            phase= row['Phase Connection'].strip()

            qcd_rec = QCD_MAP.get(sse, {})
            qcd_dt  = qcd_rec.get('qcd', None)
            cohort_offer = qcd_rec.get('offer', '') or offer_raw
            cohort = assign_cohort(qcd_dt, cohort_offer)

            project_map[sse] = {
                'id':sse,'c':city,'s':state,'o':offer,'ph':phase,
                'kw':kw,'rev':round(rev,2),'dt':d.strftime('%Y-%m-%d') if d else '',
                'qcd': qcd_dt.strftime('%Y-%m-%d') if qcd_dt else '',
                'cohort': cohort,
                'mod':0,'inv':0,'prf':0,'cab':0,'ick':0,'con':0,'ear':0,'jbx':0,
                'tsh':0,'saf':0,'ica':0,'wel':0,'ssn':0,'ebo':0,'dlg':0,'mtr':0,'wkt':0,'lad':0,
                'mt':'','mq':0,'it':'','iq':0,'mlc':0,
            }

        if cat == 'EXCLUDE':
            excluded_count += 1
            continue
        k2 = CAT_KEY.get(cat)
        if k2:
            project_map[sse][k2] = round(project_map[sse][k2] + amt, 2)
        elif cat:
            unmapped_cats[cat] += 1

        p = project_map[sse]
        if cat == 'Module' and item_name:
            if not p['mt']: p['mt'] = item_name; p['mq'] = qty
            elif p['mt'] == item_name: p['mq'] += qty
            _iname = item_name.strip()
            _lc_wp = 0
            if _iname == '540 Wp Mono Bifacial DCR-PREMIER':
                if rate != 11055:
                    _lc_wp = 540
            elif _iname == '610 Wp Mono Bifacial Topcon N-Type DCR-PREMIER':
                if rate <= 13041.8:
                    _lc_wp = 610
            if _lc_wp > 0:
                try:
                    _dt = datetime.strptime(p['dt'], '%Y-%m-%d')
                    _factor = 0.50 if (_dt.year > 2026 or (_dt.year == 2026 and _dt.month >= 7)) else 0.75
                except Exception:
                    _factor = 0.75
                p['mlc'] = round(p.get('mlc', 0) + _lc_wp * qty * _factor, 2)
        if cat == 'Inverter' and item_name:
            if not p['it']:
                p['it'] = item_name; p['iq'] = qty
            if '_inv_phase' not in p or not p['_inv_phase']:
                detected = detect_inverter_phase(item_name)
                if detected:
                    p['_inv_phase'] = detected
            inv_type = detect_inverter_type(item_name)
            if inv_type:
                proj_inv_types[sse][inv_type]['qty'] += qty
                proj_inv_types[sse][inv_type]['amt'] += amt

        if cat in ('Prefab MMS', 'Welded MMS', 'Tin Shed MMS') and item_subcat:
            short_name = shorten_mms_item_name(item_name, item_subcat)
            proj_mms_items[sse][item_subcat][short_name]['qty'] += qty
            proj_mms_items[sse][item_subcat][short_name]['amt'] += amt
            proj_mms_items[sse][item_subcat][short_name]['uom'] = uom

        if cat == 'Cables' and item_subcat:
            short_cable = shorten_cable_subcat(item_subcat)
            proj_cable_items[sse][short_cable]['qty'] += qty
            proj_cable_items[sse][short_cable]['amt'] += amt
            proj_cable_items[sse][short_cable]['cases'] += 1

print(f"Built {len(project_map):,} projects")
print(f"  Excluded rows (dongles, Safety Lifeline, Civil Work): {excluded_count:,}")

if unmapped_cells:
    print(f"\n⚠  WARNING: {len(unmapped_cells)} unmapped cell names:")
    for cell, cnt in sorted(unmapped_cells.items(), key=lambda x: -x[1]):
        print(f"    {cell}: {cnt} projects")

if unmapped_cats:
    print(f"\n⚠  WARNING: {len(unmapped_cats)} categories not in CAT_KEY:")
    for cat, cnt in sorted(unmapped_cats.items(), key=lambda x: -x[1])[:20]:
        print(f"    '{cat}': {cnt:,} rows")

if erp_mms_overrides:
    total_lost = sum(d['amt'] for d in erp_mms_overrides.values())
    print(f"\n⚠  ERP OVERRIDE ALERT: {len(erp_mms_overrides)} item_code(s) reclassified from Prefab MMS")
    print(f"   Total: ₹{total_lost:,.0f}  ({total_lost/1e7:.4f} Cr)")
    for code, d in sorted(erp_mms_overrides.items(), key=lambda x: -x[1]['amt']):
        print(f"     {code:<20}  → ERP='{d['resolved_cat']}'  ₹{d['amt']:>10,.0f}  ({len(d['sses'])} projects)")
else:
    print("\n✓ No ERP overrides reclassify Prefab MMS")

# ── Backend metering injection ─────────────────────────────────────────────────
month_metering = defaultdict(float)
no_rate_cities = defaultdict(int)
phase_mismatch_count = 0

for sse, p in project_map.items():
    inv_phase = p.get('_inv_phase', p['ph'])
    sanction_phase = p['ph']
    if inv_phase != sanction_phase:
        phase_mismatch_count += 1

    backend = calc_metering_backend(p['c'], inv_phase, sanction_phase)
    dn = dn_metering.get(sse, 0)
    total_mtr = backend + dn

    if total_mtr > 0:
        p['mtr'] = round(p['mtr'] + total_mtr, 2)

    if p['dt']:
        mkey = p['dt'][:7]
        month_metering[mkey] += total_mtr

    if p['c'] and p['c'] not in NM_RATES and backend == 0:
        no_rate_cities[p['c']] += 1

print()
print(f"  Phase mismatches: {phase_mismatch_count}")
for mkey in sorted(month_metering):
    if month_metering[mkey] > 0:
        count = sum(1 for p in project_map.values() if p['dt'].startswith(mkey))
        print(f"  Metering {mkey}: ₹{month_metering[mkey]:,.0f} → {count} projects")

if no_rate_cities:
    print(f"\n⚠  Cities not in rate table (0 metering):")
    for c, cnt in sorted(no_rate_cities.items(), key=lambda x: -x[1]):
        print(f"    {c}: {cnt} projects")

# ── Compute final COGS ────────────────────────────────────────────────────────
projects = []
for sse, p in project_map.items():
    p.pop('_inv_phase', None)
    cogs = round(p['mod']+p['inv']+p['prf']+p['cab']+p['ick']+p['con']+p['ear']+
                 p['jbx']+p['tsh']+p['saf']+p['ica']+p['wel']+p['ssn']+p['ebo']+
                 p['dlg']+p['mtr']+p['wkt']+p['lad']+p.get('mlc',0), 2)

    out = {**p, 'cogs': cogs}

    onm_val = proj_onm_amt.get(sse, 0)
    qhse_val = proj_qhse_amt.get(sse, 0)
    if onm_val: out['onm'] = round(onm_val, 2)
    if qhse_val: out['qhs'] = round(qhse_val, 2)

    ivt = proj_inv_types.get(sse)
    if ivt:
        out['ivt'] = {t: {'q': round(d['qty'],1), 'a': round(d['amt'],2)} for t, d in ivt.items()}

    mms = proj_mms_items.get(sse)
    if mms:
        out['msd'] = {}
        for subcat, items in mms.items():
            out['msd'][subcat] = {nm: {'q': round(d['qty'],2), 'a': round(d['amt'],2)} for nm, d in items.items()}

    cab_items = proj_cable_items.get(sse)
    if cab_items:
        out['cbd'] = {sc: {'q': round(d['qty'],2), 'a': round(d['amt'],2), 'n': d['cases']} for sc, d in cab_items.items()}

    projects.append(out)

# ── Monthly ONM & QHSE totals ──────────────────────────────────────────────────
monthly_onm_qhse = defaultdict(lambda: {'onm': 0, 'qhs': 0})

with gzip.open('data.csv.gz', 'rt', encoding='utf-8', errors='replace') as f:
    reader = csv.DictReader(f)
    for row in reader:
        # ── CHANGED: parent renamed to parent_id ──
        parent = row.get('parent_id', '').strip()
        posting_date_str = row.get('posting_date', '').strip()
        amount = float(row.get('amount', 0) or 0)

        if not posting_date_str or not parent:
            continue

        try:
            date_obj = datetime.strptime(posting_date_str, '%Y-%m-%d')
            month_key = f"{date_obj.year}-{date_obj.month:02d}"

            if parent.lower().startswith('onm'):
                monthly_onm_qhse[month_key]['onm'] += amount
            elif parent.lower().startswith('qhse'):
                monthly_onm_qhse[month_key]['qhs'] += amount
        except:
            pass

metadata = {
    'monthly_onm_qhse': {k: {'onm': round(v['onm'], 2), 'qhs': round(v['qhs'], 2)} for k, v in monthly_onm_qhse.items()}
}

# ── Write output ──────────────────────────────────────────────────────────────
output = {'_meta': metadata, 'projects': projects}
json_str = json.dumps(output, separators=(',',':'))

import shutil
with open('projects_temp.json', 'w', encoding='utf-8') as f:
    f.write(json_str)
with open('projects_temp.json', 'rb') as f_in, \
     gzip.open('projects.json.gz', 'wb', compresslevel=6) as f_out:
    shutil.copyfileobj(f_in, f_out)
try:
    os.remove('projects_temp.json')
except Exception:
    pass

raw_mb = len(json_str)/1e6
gz_mb  = os.path.getsize('projects.json.gz')/1e6
print(f"\nOutput: {len(projects):,} projects | JSON {raw_mb:.1f} MB → gz {gz_mb:.2f} MB")

# ── Quick verification ─────────────────────────────────────────────────────────
print("\n── Verification ──")

for mo, label in [(1, 'Jan 26'), (2, 'Feb 26'), (3, 'Mar 26')]:
    ps = [p for p in projects if p['dt'].startswith(f'2026-{mo:02d}')]
    if not ps: continue
    rev  = sum(p['rev']  for p in ps)
    cogs = sum(p['cogs'] for p in ps)
    mtr  = sum(p['mtr']  for p in ps)
    gm   = (rev-cogs)/rev*100 if rev else 0
    print(f"\n  {label}: {len(ps)} projects | Rev={rev/1e7:.2f}Cr | COGS={cogs/1e7:.2f}Cr | GM%={gm:.2f}%")
    for cat_name, key in sorted(CAT_KEY.items(), key=lambda x: -sum(p[x[1]] for p in ps if x[1] in p)):
        total = sum(p.get(key, 0) for p in ps)
        if total > 0:
            print(f"    {cat_name:25s}: ₹{total/1e7:.2f}Cr")

print("\n── Metering Accuracy ──")
for mo, label, actual_mtr in [(1, 'Jan 26', 5926077), (2, 'Feb 26', 5755707), (3, 'Mar 26', 7909163)]:
    ps = [p for p in projects if p['dt'].startswith(f'2026-{mo:02d}')]
    mtr = sum(p['mtr'] for p in ps)
    delta = mtr - actual_mtr
    pct   = delta / actual_mtr * 100 if actual_mtr else 0
    print(f"  {label}: Metering={mtr:,.0f} (actual {actual_mtr:,.0f}, delta {delta:+,.0f} = {pct:+.2f}%)")

print("\n── Cohort Assignment Coverage ──")
total_p = len(projects)
with_cohort = sum(1 for p in projects if p.get('cohort',''))
with_qcd    = sum(1 for p in projects if p.get('qcd',''))
print(f"  Projects with QCD date : {with_qcd:,} / {total_p:,} ({with_qcd/total_p*100:.1f}%)")
print(f"  Projects with cohort   : {with_cohort:,} / {total_p:,} ({with_cohort/total_p*100:.1f}%)")
if with_cohort < total_p:
    missing = total_p - with_cohort
    print(f"  ⚠ {missing:,} projects have no cohort")

print("\nDone.")
